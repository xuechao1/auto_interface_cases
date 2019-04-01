__author__ = 'Emerson'
# 把测试数据从Excel里面获取出来
from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    # 读取未注册手机号
    def no_reg_tel(self):
        wb = load_workbook(self.file_path)
        sheet = wb['init']
        no_reg_tel = sheet.cell(1, 2).value
        return no_reg_tel  # 从Excel里面获取到未注册的手机号

    def read_data(self):
        no_reg_tel = self.no_reg_tel()
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_data = []  # 存储所有行的数据
        for i in range(2, sheet.max_row + 1):
            sub_data = []  # 存储每一行的数据
            for j in range(1, 8):
                if j == 6:
                    param = eval(sheet.cell(i, 6).value)
                    if param['mobilephone'] == 'first_tel':
                        param['mobilephone'] = no_reg_tel
                        sub_data.append(param)
                else:
                    sub_data.append(sheet.cell(i, j).value)
            test_data.append(sub_data)
            self.update_tel(str(int(no_reg_tel) + 1))
        return test_data

    def update_tel(self, new_tel):
        wb = load_workbook(self.file_path)
        sheet = wb['init']
        sheet.cell(1, 2).value = new_tel
        wb.save(self.file_path)

    def write_data(self, row, actually, result):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        sheet.cell(row, 8).value = actually
        sheet.cell(row, 9).value = result
        wb.save(self.file_path)


if __name__ == '__main__':
    test_data = DoExcel('test_case.xlsx', 'test_data').read_data()
    print(test_data)
