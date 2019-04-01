__author__ = 'Emerson'
# 把测试数据从Excel里面获取出来
from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    # 读取未注册手机号
    def no_reg_tel(self):
        wb = load_workbook(self.file_path)
        sheet = wb['init']
        no_reg_tel = sheet.cell(1, 2).value
        return no_reg_tel  # 从Excel里面获取到未注册的手机号

    def read_data(self, mode, case_list):
        no_reg_tel = self.no_reg_tel()
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        test_data = []  # 存储所有行的数据
        if mode == '1':  # 执行所有的用例 就要加载所有的用例
            for i in range(2, sheet.max_row + 1):
                sub_data = []  # 存储每一行的数据or存到一个字典里面sub_data={}
                for j in range(1, 9):
                    if j == 6:
                        param = eval(sheet.cell(i, 6).value)
                        if param['mobilephone'] == 'first_tel':
                            param['mobilephone'] = no_reg_tel  # 变换替换
                        sub_data.append(param)
                    elif j == 8:
                        check_sql = eval(sheet.cell(i, 8).value)
                        if check_sql['sql_data'] == 'first_tel':
                            check_sql['sql_data'] = no_reg_tel  # 变换替换
                        sub_data.append(check_sql)
                    else:
                        sub_data.append(sheet.cell(i, j).value)
                test_data.append(sub_data)
        elif mode == '0':
            for i in case_list:
                sub_data = []  # 存储每一行的数据
                for j in range(1, 9):
                    if j == 6:
                        param = eval(sheet.cell(i + 1, 6).value)
                        if param['mobilephone'] == 'first_tel':
                            param['mobilephone'] = no_reg_tel
                        sub_data.append(param)
                    if j == 8:
                        check_sql = eval(sheet.cell(i, 8).value)
                        if check_sql['sql_data'] == 'first_tel':
                            check_sql['sql_data'] = no_reg_tel  # 变换替换
                        sub_data.append(check_sql)
                    else:
                        sub_data.append(sheet.cell(i + 1, j).value)
                test_data.append(sub_data)
        self.update_tel(str(int(no_reg_tel) + 1))
        return test_data

    def update_tel(self, new_tel):
        wb = load_workbook(self.file_path)
        sheet = wb['init']
        sheet.cell(1, 2).value = new_tel
        wb.save(self.file_path)

    def write_data(self, row, mode, actually, result):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        if mode == 1:
            sheet.cell(row, 9).value = actually
            sheet.cell(row, 10).value = result
        elif mode == 2:
            sheet.cell(row, 11).value = actually
            sheet.cell(row, 12).value = result
        wb.save(self.file_path)


if __name__ == '__main__':
    from conf import project_path

    test_data = DoExcel(project_path.test_data_path, 'test_data').read_data('1', [2, 3])
    print(test_data)
